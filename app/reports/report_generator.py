import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


class ReporteErroresExcel:
    def __init__(self, df_original, errores, ruta_config=None):

        # Cargar configuración desde YAML
        if ruta_config is None:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            ruta_config = os.path.join(base_dir, "config/settings.yaml")

        with open(ruta_config, "r", encoding="utf-8") as file:
            config = yaml.safe_load(file)

        conf_reporte = config.get("reporte_errores", {})

        self.df = df_original.copy()
        self.errores = errores
        self.color_error_total = conf_reporte.get("color_error_total")
        self.color_error_campo = conf_reporte.get("color_error_campo")
        self.nombre_columna_errores = conf_reporte.get("nombre_columna_errores")
        self.ruta_salida = conf_reporte.get("ruta_salida")

    def aplicar_errores(self):
        """Agrega columna de errores al DataFrame."""
        errores_dict = {e['fila']: "\n".join(e['errores']) for e in self.errores}
        self.df[self.nombre_columna_errores] = [
            errores_dict.get(i + 2, "") for i in range(len(self.df))
        ]
        return self.df

    def _resaltar_errores(self, ws, columnas_dict, col_errores):
        """Aplica colores a las celdas con errores."""
        fill_total = PatternFill(start_color=self.color_error_total, end_color=self.color_error_total,
                                 fill_type="solid")
        fill_campo = PatternFill(start_color=self.color_error_campo, end_color=self.color_error_campo,
                                 fill_type="solid")

        for error_info in self.errores:
            fila = error_info["fila"]
            errores_texto = error_info["errores"]

            # Resaltar columna de errores
            ws.cell(row=fila, column=col_errores).fill = fill_total

            # Detectar campos mencionados en los errores
            for error_msg in errores_texto:
                for nombre_col, idx in columnas_dict.items():
                    if nombre_col.lower() in error_msg.lower():
                        ws.cell(row=fila, column=idx).fill = fill_campo

    def exportar(self, ruta_salida=None, writer=None, nombre_hoja=None):
        if ruta_salida is None:
            ruta_salida = self.ruta_salida

        df_con_errores = self.aplicar_errores()

        if writer is None:
            # Guardar en archivo independiente
            df_con_errores.to_excel(ruta_salida, index=False)
            wb = load_workbook(ruta_salida)
            ws = wb.active
            ws.freeze_panes = "A2"
        else:
            # Guardar en hoja del archivo abierto con ExcelWriter
            if nombre_hoja is None:
                nombre_hoja = "Hoja1"
            df_con_errores.to_excel(writer, index=False, sheet_name=nombre_hoja)
            wb = writer.book
            ws = wb[nombre_hoja]
            ws.freeze_panes = "A2"

        columnas_dict = {col: idx + 1 for idx, col in enumerate(df_con_errores.columns)}
        col_errores = columnas_dict.get(self.nombre_columna_errores)

        if col_errores:
            self._resaltar_errores(ws, columnas_dict, col_errores)

        if writer is None:
            wb.save(ruta_salida)
            print(f"🔴 Reporte generado con errores resaltados: {ruta_salida}")
        else:
            # No guardamos porque lo hace la clase que abrió el writer
            print(f"🔴 Hoja '{nombre_hoja}' generada con errores resaltados.")

